import os
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog
from openpyxl import Workbook
from openpyxl import load_workbook

class App(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.title = 'Find Doubles'
        self.left : int = 500
        self.top : int = 500
        self.height : int = 250
        self.width : int = 250
        self.initUI()
        


    def initUI(self):
        self.statusBar().showMessage('Select excel')
        browse_button = QPushButton('Browse', self)
        browse_button.clicked.connect(self.browse_file)
        browse_button.move(50,25)
        self.show()

    def browse_file(self):
        desktop = os.path.expanduser("~\Desktop\\") #path for current user desktop
        
        fileName, _ = QFileDialog.getOpenFileName(self,"Select Excel", "","Excel Files(*.xlsx)") #Grab File
        xl_to_check = Workbook()
        duplicate_xl = Workbook()
        xl_to_check = load_workbook(fileName)
        xl_to_check_sheet = xl_to_check.active
        # make a new list to store the IDs
        xl_duplicates = []
        xl_column_to_check = []
        for id in xl_to_check_sheet.iter_rows(min_row=2, max_row=xl_to_check_sheet.max_row, min_col=1, max_col=1,values_only=True):
           
            for id_to_check in xl_to_check_sheet.iter_rows(min_row=1, max_row=xl_to_check_sheet.max_row, min_col=3, max_col=3,values_only=True):
                if id == id_to_check:
                    xl_duplicates.append(id_to_check)
        duplicate_xl.create_sheet("Duplicates")
        duplicate_xl.active.append(xl_duplicates)
        duplicate_xl.save(desktop + 'Duplicates.xlsx')
        xl_to_check.close()
        self.statusBar().showMessage('Process Finished')
        
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    ex = App()
    sys.exit(app.exec_())